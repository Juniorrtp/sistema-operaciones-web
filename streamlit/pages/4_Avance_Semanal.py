import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from utils.styles import apply_custom_styles
from utils.api_client import (
    fetch_from_api,
    load_movimientos_general,
    load_movimientos_detalles,
    load_metros_general,
    load_metros_detalles,
    load_stock_from_api,
    load_objetivos
    
)
# Aplicar estilos personalizados
apply_custom_styles()
# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)



# ============================================
# FUNCIÓN PARA CREAR TABLAS HTML PERSONALIZADAS
# ============================================

def crear_tabla_html(df, titulo=None):
    """Crea una tabla HTML personalizada sin scroll y con estilo profesional"""
    
    if df.empty:
        return '<p style="text-align:center; color:#6c757d; padding:20px;">ℹ️ No hay datos para mostrar</p>'
    
    # Obtener columnas
    columnas = df.columns.tolist()
    
    # Construir HTML de la tabla
    html = ""
    
    if titulo:
        html += f'<p style="font-size:14px; font-weight:600; color:#2c3e50; margin:10px 0 5px 0;">{titulo}</p>'
    
    html += '<div style="display:flex; justify-content:center; width:100%; overflow:visible;">'
    html += '<table style="border-collapse:collapse; border:2px solid #2c3e50; font-family:Segoe UI, Arial, sans-serif; font-size:13px; min-width:400px; max-width:100%; margin:0 auto;">'
    
    # Encabezados
    html += '<thead>'
    html += '<tr style="background:linear-gradient(135deg, #2c3e50 0%, #34495e 100%);">'
    for col in columnas:
        html += f'<th style="color:#ffffff; font-weight:bold; font-size:12px; text-align:center; padding:8px 14px; border:1px solid #1a252f; text-transform:uppercase; letter-spacing:0.5px; white-space:nowrap;">{col}</th>'
    html += '</tr>'
    html += '</thead>'
    
    # Cuerpo
    html += '<tbody>'
    for idx, row in df.iterrows():
        # Fila alterna
        bg_color = '#f8f9fa' if idx % 2 == 1 else '#ffffff'
        html += f'<tr style="background-color:{bg_color};">'
        
        for i, col in enumerate(columnas):
            valor = row[col]
            
            # Estilo según tipo de columna
            if col == 'Codigo':
                estilo = 'text-align:center; font-weight:600; color:#1a5276; white-space:nowrap;'
            elif col == 'Descripcion':
                estilo = 'text-align:left; min-width:120px; max-width:250px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;'
            elif col == 'Stock' and valor != '':
                estilo = 'text-align:center; font-weight:bold; color:#1a5276; background-color:#ebf5fb;'
            elif isinstance(valor, (int, float)):
                estilo = 'text-align:right;'
            else:
                estilo = 'text-align:left;'
            
            # Formatear número
            if isinstance(valor, (int, float)):
                if col == 'Stock' and valor == '':
                    valor_display = ''
                else:
                    valor_display = f'{valor:,.0f}'
            else:
                valor_display = valor
            
            html += f'<td style="padding:6px 12px; border:1px solid #bdc3c7; color:#2c3e50; font-size:12px; {estilo}">{valor_display}</td>'
        
        html += '</tr>'
    html += '</tbody>'
    
    html += '</table>'
    html += '</div>'
    
    return html

# ============================================
# FUNCIONES DE CONEXIÓN A API
# ============================================



# ============================================
# CARGA DE DATOS CON CACHÉ
# ============================================

@st.cache_data(ttl=300)
def load_movimientos_general(fecha_desde=None, fecha_hasta=None):
    params = {"limit": 5000}
    if fecha_desde:
        params["fecha_desde"] = fecha_desde
    if fecha_hasta:
        params["fecha_hasta"] = fecha_hasta
    
    data = fetch_from_api("movimientos", params)
    
    if data:
        for row in data:
            if 'compania' in row and row['compania']:
                row['compania'] = row['compania'].strip()
            if 'tipo_perforacion' in row and row['tipo_perforacion']:
                row['tipo_perforacion'] = row['tipo_perforacion'].strip()
            if 'mes' in row and row['mes']:
                row['mes'] = row['mes'].strip().upper()
            if 'movimiento' in row and row['movimiento']:
                row['movimiento'] = row['movimiento'].strip().upper()
            if 'estado' in row and row['estado']:
                row['estado'] = row['estado'].strip().upper()
    return data

@st.cache_data(ttl=300)
def load_movimientos_detalles():
    return fetch_from_api("detalles-movimientos")

@st.cache_data(ttl=300)
def load_metros_general(fecha_desde=None, fecha_hasta=None):
    params = {"limit": 5000}
    if fecha_desde:
        params["fecha_desde"] = fecha_desde
    if fecha_hasta:
        params["fecha_hasta"] = fecha_hasta
    
    data = fetch_from_api("metros", params)
    
    if data:
        for row in data:
            if 'compania' in row and row['compania']:
                row['compania'] = row['compania'].strip()
            if 'tipo_perforacion' in row and row['tipo_perforacion']:
                row['tipo_perforacion'] = row['tipo_perforacion'].strip()
            if 'mes' in row and row['mes']:
                row['mes'] = row['mes'].strip().upper()
    return data

@st.cache_data(ttl=300)
def load_metros_detalles():
    return fetch_from_api("metros-detalles")

@st.cache_data(ttl=600)
def load_stock_from_api():
    return fetch_from_api("stock")


# ============================================
# FUNCIÓN: RENDIMIENTO DE ACEROS (desde Rendimiento.py)
# ============================================

@st.cache_data(ttl=300)
def process_rendimiento_aceros(fecha_desde, fecha_hasta, año=None, mes=None, compania="TODAS"):
    """Procesa datos de rendimiento de aceros (misma lógica que Rendimiento)"""
    
    familias_target = ['SHANK', 'ACOPLES', 'BARRAS', 'RIMADORAS']
    
    # Cargar datos
    mov_detalles = load_movimientos_detalles()
    met_detalles = load_metros_detalles()
    mov_general = load_movimientos_general(fecha_desde, fecha_hasta)
    objetivos = load_objetivos()
    
    if not mov_general or not mov_detalles:
        return {}
    
    df_mov_gen = pd.DataFrame(mov_general)
    df_mov_det = pd.DataFrame(mov_detalles)
    df_met_det = pd.DataFrame(met_detalles)
    
    # Filtrar movimientos generales
    if año and mes:
        df_mov_gen = df_mov_gen[
            (df_mov_gen['ano'] == int(año)) &
            (df_mov_gen['mes'] == mes.upper())
        ]
    
    if fecha_desde and fecha_hasta:
        df_mov_gen = df_mov_gen[
            (pd.to_datetime(df_mov_gen['fecha']) >= pd.to_datetime(fecha_desde)) &
            (pd.to_datetime(df_mov_gen['fecha']) <= pd.to_datetime(fecha_hasta))
        ]
    
    if compania != "TODAS":
        df_mov_gen = df_mov_gen[df_mov_gen['compania'] == compania.strip()]
    
    if df_mov_gen.empty:
        return {}
    
    mov_ids = df_mov_gen['id'].tolist()
    df_mov_det_filtrado = df_mov_det[df_mov_det['entrega_id'].isin(mov_ids)]
    
    if df_mov_det_filtrado.empty:
        return {}
    
    # Unir con generales para obtener equipo y tipo_perforacion
    df_mov_det_filtrado = df_mov_det_filtrado.merge(
        df_mov_gen[['id', 'equipo', 'tipo_perforacion']],
        left_on='entrega_id',
        right_on='id',
        how='left'
    )
    
    # Filtrar solo familias target
    df_mov_det_filtrado = df_mov_det_filtrado[
        df_mov_det_filtrado['familia'].str.upper().isin(familias_target)
    ]
    
    if df_mov_det_filtrado.empty:
        return {}
    
    # Crear diccionario de objetivos
    obj_dict = {}
    for obj in objetivos:
        tipo = obj.get('Tipo Perforacion', '')
        familia = obj.get('Acero', '')
        objetivo_val = obj.get('Objetivo', 0)
        if objetivo_val is None:
            objetivo_val = 0
        obj_dict[(tipo, familia)] = objetivo_val
    
    # Obtener metros por equipo y familia
    df_met_gen = pd.DataFrame(load_metros_general())
    
    if año and mes:
        df_met_gen = df_met_gen[
            (df_met_gen['ano'] == int(año)) &
            (df_met_gen['mes'] == mes.upper())
        ]
    
    if fecha_desde and fecha_hasta:
        df_met_gen = df_met_gen[
            (pd.to_datetime(df_met_gen['fecha']) >= pd.to_datetime(fecha_desde)) &
            (pd.to_datetime(df_met_gen['fecha']) <= pd.to_datetime(fecha_hasta))
        ]
    
    if compania != "TODAS":
        df_met_gen = df_met_gen[df_met_gen['compania'] == compania.strip()]
    
    met_ids = df_met_gen['id'].tolist()
    df_met_det_filtrado = df_met_det[df_met_det['registro_id'].isin(met_ids)]
    
    # Agrupar por tipo_perforacion, familia
    agrupado = df_mov_det_filtrado.groupby(['tipo_perforacion', 'familia']).agg({
        'cantidad': lambda x: x.abs().sum()
    }).reset_index()
    
    # Calcular resultados por tipo y familia
    resultados = []
    
    for tipo in agrupado['tipo_perforacion'].unique():
        df_tipo = agrupado[agrupado['tipo_perforacion'] == tipo]
        
        # Obtener metros para este tipo
        met_ids_tipo = df_met_gen[df_met_gen['tipo_perforacion'] == tipo]['id'].tolist()
        df_met_tipo = df_met_det_filtrado[df_met_det_filtrado['registro_id'].isin(met_ids_tipo)]
        
        total_mp_tipo = df_met_tipo['total_mp'].sum()
        mp_rimado_tipo = df_met_tipo['mp_rimado'].sum()
        
        for _, row in df_tipo.iterrows():
            familia = row['familia']
            cantidad = row['cantidad']
            
            if familia.upper() == 'RIMADORAS':
                metros = mp_rimado_tipo
            else:
                metros = total_mp_tipo
            
            rendimiento = metros / cantidad if cantidad > 0 else 0
            objetivo = obj_dict.get((tipo, familia), 0)
            eficiencia = (rendimiento / objetivo * 100) if objetivo > 0 else 0
            
            resultados.append({
                'Tipo_Perforacion': tipo,
                'Familia': familia,
                'Cantidad': cantidad,
                'Metros': metros,
                'Rendimiento': rendimiento,
                'Objetivo': objetivo,
                'Eficiencia': eficiencia
            })
    
    if not resultados:
        return {}
    
    df_resultado = pd.DataFrame(resultados)
    
    # Agrupar por tipo_perforacion
    tipos = df_resultado['Tipo_Perforacion'].unique()
    resultados_final = {}
    
    for tipo in tipos:
        resultados_final[tipo] = df_resultado[df_resultado['Tipo_Perforacion'] == tipo].drop(columns=['Tipo_Perforacion'])
    
    return resultados_final
# ============================================
# FUNCIÓN: PROCESAR CONSUMOS
# ============================================

@st.cache_data(ttl=300)
def process_consumos(fecha_desde, fecha_hasta, año=None, mes=None):
    """Procesa consumos por estado -> tipo_perforacion -> compañía"""
    
    logger.info("🔍 PROCESANDO CONSUMOS")
    
    mov_detalles = load_movimientos_detalles()
    mov_general = load_movimientos_general(fecha_desde, fecha_hasta)
    
    if not mov_general or not mov_detalles:
        return {}
    
    df_mov_gen = pd.DataFrame(mov_general)
    df_mov_det = pd.DataFrame(mov_detalles)
    
    if año and mes:
        df_mov_gen = df_mov_gen[
            (df_mov_gen['ano'] == int(año)) &
            (df_mov_gen['mes'] == mes.strip().upper())
        ]
    
    if fecha_desde and fecha_hasta:
        df_mov_gen = df_mov_gen[
            (pd.to_datetime(df_mov_gen['fecha']) >= pd.to_datetime(fecha_desde)) &
            (pd.to_datetime(df_mov_gen['fecha']) <= pd.to_datetime(fecha_hasta))
        ]
    
    if df_mov_gen.empty:
        return {}
    
    mov_ids = df_mov_gen['id'].tolist()
    df_mov_det_filtrado = df_mov_det[df_mov_det['entrega_id'].isin(mov_ids)]
    
    if df_mov_det_filtrado.empty:
        return {}
    
    df_mov_det_filtrado = df_mov_det_filtrado.merge(
        df_mov_gen[['id', 'estado', 'compania']],
        left_on='entrega_id',
        right_on='id',
        how='left'
    )
    
    df_mov_det_filtrado['cantidad_abs'] = df_mov_det_filtrado['cantidad'].abs()
    
    stock_data = load_stock_from_api()
    stock_dict = {}
    for item in stock_data:
        stock_dict[item['codigo']] = item['stock']
    
    estados = df_mov_det_filtrado['estado'].dropna().unique()
    tipos_perf = df_mov_det_filtrado['tipo_perforacion'].dropna().unique()
    
    resultados = {}
    codigos_mostrados = set()
    
    for estado in sorted(estados):
        df_estado = df_mov_det_filtrado[df_mov_det_filtrado['estado'] == estado]
        resultados[estado] = {}
        
        for tipo in sorted(tipos_perf):
            df_tipo = df_estado[df_estado['tipo_perforacion'] == tipo]
            
            if df_tipo.empty:
                continue
            
            agrupado = df_tipo.groupby(['codigo', 'descripcion']).agg({
                'cantidad_abs': 'sum',
                'compania': lambda x: x.tolist()
            }).reset_index()
            
            resultados_tipo = []
            
            for _, row in agrupado.iterrows():
                codigo = row['codigo']
                descripcion = row['descripcion']
                cantidad_total = row['cantidad_abs']
                
                companias_consumo = {}
                df_companias = df_tipo[df_tipo['codigo'] == codigo]
                for compania in df_companias['compania'].unique():
                    companias_consumo[compania] = df_companias[df_companias['compania'] == compania]['cantidad_abs'].sum()
                
                if codigo not in codigos_mostrados:
                    stock = stock_dict.get(codigo, 0)
                    codigos_mostrados.add(codigo)
                else:
                    stock = None
                
                fila = {'Codigo': codigo, 'Descripcion': descripcion}
                
                for compania in sorted(df_tipo['compania'].unique()):
                    fila[compania] = companias_consumo.get(compania, 0)
                
                fila['Total'] = cantidad_total
                fila['Stock'] = stock if stock is not None else ''
                
                resultados_tipo.append(fila)
            
            if resultados_tipo:
                resultados[estado][tipo] = pd.DataFrame(resultados_tipo)
    
    return resultados

# ============================================
# FUNCIÓN: PROCESAR METROS
# ============================================

@st.cache_data(ttl=300)
def process_metros(fecha_desde, fecha_hasta, año=None, mes=None):
    """Procesa metros por tipo_perforacion"""
    
    logger.info("🔍 PROCESANDO METROS")
    
    met_detalles = load_metros_detalles()
    met_general = load_metros_general(fecha_desde, fecha_hasta)
    
    if not met_general or not met_detalles:
        return pd.DataFrame()
    
    df_met_gen = pd.DataFrame(met_general)
    df_met_det = pd.DataFrame(met_detalles)
    
    if año and mes:
        df_met_gen = df_met_gen[
            (df_met_gen['ano'] == int(año)) &
            (df_met_gen['mes'] == mes.strip().upper())
        ]
    
    if fecha_desde and fecha_hasta:
        df_met_gen = df_met_gen[
            (pd.to_datetime(df_met_gen['fecha']) >= pd.to_datetime(fecha_desde)) &
            (pd.to_datetime(df_met_gen['fecha']) <= pd.to_datetime(fecha_hasta))
        ]
    
    if df_met_gen.empty:
        return pd.DataFrame()
    
    met_ids = df_met_gen['id'].tolist()
    df_met_det_filtrado = df_met_det[df_met_det['registro_id'].isin(met_ids)]
    
    if df_met_det_filtrado.empty:
        return pd.DataFrame()
    
    df_met_det_filtrado = df_met_det_filtrado.merge(
        df_met_gen[['id', 'tipo_perforacion']],
        left_on='registro_id',
        right_on='id',
        how='left'
    )
    
    resultado = df_met_det_filtrado.groupby('tipo_perforacion').agg({
        'total_mp': 'sum',
        'mp_rimado': 'sum'
    }).reset_index()
    
    resultado = resultado.rename(columns={
        'tipo_perforacion': 'Tipo Perforación',
        'total_mp': 'MP Total',
        'mp_rimado': 'MP Rimado'
    })
    
    return resultado

# ============================================
# FUNCIÓN: EXPORTAR A EXCEL
# ============================================

def export_to_excel(consumos_data, df_metros, fecha_inicio, fecha_fin):
    output = io.BytesIO()
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws_consumos = wb.active
    ws_consumos.title = "Consumos"
    
    row_num = 1
    ws_consumos['A1'] = f"Reporte de Consumos - {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
    ws_consumos['A1'].font = Font(bold=True, size=14)
    ws_consumos.merge_cells('A1:H1')
    row_num = 3
    
    for estado in sorted(consumos_data.keys()):
        ws_consumos[f'A{row_num}'] = f"ESTADO: {estado}"
        ws_consumos[f'A{row_num}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_consumos[f'A{row_num}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_consumos.merge_cells(f'A{row_num}:H{row_num}')
        row_num += 1
        
        for tipo in sorted(consumos_data[estado].keys()):
            df = consumos_data[estado][tipo]
            
            if df.empty:
                continue
            
            ws_consumos[f'A{row_num}'] = f"Tipo: {tipo}"
            ws_consumos[f'A{row_num}'].font = Font(bold=True, size=11)
            row_num += 1
            
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws_consumos.cell(row=row_num, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row_num += 1
            
            for _, row in df.iterrows():
                for col_idx, header in enumerate(headers, 1):
                    value = row[header]
                    cell = ws_consumos.cell(row=row_num, column=col_idx, value=value)
                    cell.border = border
                    
                    if header == 'Stock' and value != '':
                        cell.font = Font(bold=True)
                    
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left')
                
                row_num += 1
            
            row_num += 1
    
    if not df_metros.empty:
        ws_metros = wb.create_sheet("Metros")
        
        ws_metros['A1'] = f"Resumen de Metros - {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
        ws_metros['A1'].font = Font(bold=True, size=14)
        ws_metros.merge_cells('A1:C1')
        
        headers_metros = list(df_metros.columns)
        for col_idx, header in enumerate(headers_metros, 1):
            cell = ws_metros.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = 4
        for _, row in df_metros.iterrows():
            for col_idx, header in enumerate(headers_metros, 1):
                cell = ws_metros.cell(row=row_num, column=col_idx, value=row[header])
                cell.border = border
                if isinstance(row[header], (int, float)):
                    cell.alignment = Alignment(horizontal='right')
            row_num += 1
    
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
            except AttributeError:
                continue
            
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

# ============================================
# FUNCIÓN: OBTENER KPIs
# ============================================

@st.cache_data(ttl=300)
def get_kpis(fecha_desde, fecha_hasta, año=None, mes=None):
    consumos_data = process_consumos(fecha_desde, fecha_hasta, año, mes)
    
    total_consumo = 0
    total_codigos = 0
    
    for estado in consumos_data:
        for tipo in consumos_data[estado]:
            df = consumos_data[estado][tipo]
            if not df.empty and 'Total' in df.columns:
                total_consumo += df['Total'].sum()
                total_codigos += len(df)
    
    df_metros = process_metros(fecha_desde, fecha_hasta, año, mes)
    total_mp = df_metros['MP Total'].sum() if not df_metros.empty else 0
    total_rimado = df_metros['MP Rimado'].sum() if not df_metros.empty else 0
    
    tipos = df_metros['Tipo Perforación'].tolist() if not df_metros.empty else []
    
    return {
        'total_consumo': total_consumo,
        'total_codigos': total_codigos,
        'total_mp': total_mp,
        'total_rimado': total_rimado,
        'tipos_count': len(tipos)
    }

# ============================================
# ESTILOS CSS PARA IMPRESIÓN Y MÉTRICAS
# ============================================

st.markdown("""
    <style>
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        .stDeployButton {display: none;}
        
        .metric-container {
            background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
            padding: 15px 20px;
            border-radius: 12px;
            border-left: 5px solid #2c3e50;
            border-right: 1px solid #e9ecef;
            border-top: 1px solid #e9ecef;
            border-bottom: 1px solid #e9ecef;
            box-shadow: 0 4px 6px rgba(0,0,0,0.07);
            text-align: center;
        }
        
        .metric-label {
            font-size: 11px;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 1.5px;
            font-weight: 600;
            margin-bottom: 5px;
        }
        
        .metric-value {
            font-size: 26px;
            font-weight: 700;
            color: #1a252f;
            line-height: 1.2;
        }
        
        .metric-delta {
            font-size: 13px;
            color: #28a745;
            font-weight: 500;
            margin-top: 3px;
        }
        
        @media print {
            .stApp { background: white !important; padding: 20px !important; margin: 0 !important; }
            .stButton, .stSelectbox, .stDateInput { display: none !important; }
            .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 { color: #1a252f !important; }
            .stMarkdown hr { border: 2px solid #2c3e50 !important; }
            
            /* Tablas en impresión */
            table {
                page-break-inside: avoid !important;
                font-size: 10px !important;
                border: 2px solid #000 !important;
            }
            table th {
                background: #2c3e50 !important;
                color: white !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }
            table td {
                border: 1px solid #999 !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }
            table tr:nth-child(even) td {
                background-color: #f5f5f5 !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }
            .metric-container {
                border: 1px solid #333 !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }
        }
    </style>
""", unsafe_allow_html=True)

# ============================================
# FILTROS Y CARGA DE DATOS
# ============================================

st.title("📅 Avance Semanal - Reporte Gerencial")

with st.spinner("Cargando datos..."):
    movimientos_data = load_movimientos_general()
    df_mov = pd.DataFrame(movimientos_data)

if not df_mov.empty:
    años_disponibles = sorted(df_mov['ano'].unique())
    meses_disponibles = sorted(df_mov['mes'].unique())
else:
    años_disponibles = [2024, 2025, 2026]
    meses_disponibles = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                         'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

if not df_mov.empty:
    df_mov_salida = df_mov[df_mov['movimiento'] == 'SALIDA']
    if not df_mov_salida.empty:
        df_mov_salida['fecha_dt'] = pd.to_datetime(df_mov_salida['fecha'])
        ultima_fecha = df_mov_salida['fecha_dt'].max()
        ultimo_mes = ultima_fecha.strftime('%B').upper()
        ultimo_ano = ultima_fecha.year
    else:
        ultimo_mes = None
        ultimo_ano = None
else:
    ultimo_mes = None
    ultimo_ano = None

# ============================================
# FILTROS EN UNA SOLA LÍNEA
# ============================================

st.markdown("---")

col_f1, col_f2, col_f3, col_f4, col_f5, col_f6 = st.columns([1, 1, 1, 1, 0.8, 0.8])

with col_f1:
    if ultimo_ano and ultimo_ano in años_disponibles:
        idx_ano = años_disponibles.index(ultimo_ano)
    else:
        idx_ano = len(años_disponibles) - 1 if años_disponibles else 0
    
    año_seleccionado = st.selectbox(
        "Año",
        años_disponibles,
        index=idx_ano,
        key="año_avance"
    )

with col_f2:
    if ultimo_mes and ultimo_mes in meses_disponibles:
        idx_mes = meses_disponibles.index(ultimo_mes)
    else:
        idx_mes = len(meses_disponibles) - 1 if meses_disponibles else 0
    
    mes_seleccionado = st.selectbox(
        "Mes",
        meses_disponibles,
        index=idx_mes,
        key="mes_avance"
    )

with col_f3:
    fecha_inicio = st.date_input(
        "Desde",
        value=datetime.now() - timedelta(days=30),
        key="fecha_inicio"
    )

with col_f4:
    fecha_fin = st.date_input(
        "Hasta",
        value=datetime.now(),
        key="fecha_fin"
    )

with col_f5:
    if st.button("🔄 Actualizar", type="primary", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

with col_f6:
    with st.spinner("Preparando..."):
        consumos_export = process_consumos(fecha_inicio, fecha_fin, año_seleccionado, mes_seleccionado)
        metros_export = process_metros(fecha_inicio, fecha_fin, año_seleccionado, mes_seleccionado)
        
        if consumos_export or not metros_export.empty:
            excel_file = export_to_excel(consumos_export, metros_export, fecha_inicio, fecha_fin)
            st.download_button(
                label="📥 Exportar",
                data=excel_file,
                file_name=f"Reporte_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

st.markdown("---")

# ============================================
# KPIs
# ============================================

with st.spinner("Calculando indicadores..."):
    kpis = get_kpis(fecha_inicio, fecha_fin, año_seleccionado, mes_seleccionado)

col_k1, col_k2, col_k3, col_k4, col_k5 = st.columns(5)

with col_k1:
    st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">📦 Consumo Total</div>
            <div class="metric-value">{kpis['total_consumo']:,.0f}</div>
            <div class="metric-delta">{kpis['total_codigos']} códigos</div>
        </div>
    """, unsafe_allow_html=True)

with col_k2:
    st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">📏 MP Total</div>
            <div class="metric-value">{kpis['total_mp']:,.2f} m</div>
        </div>
    """, unsafe_allow_html=True)

with col_k3:
    st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">🔄 MP Rimado</div>
            <div class="metric-value">{kpis['total_rimado']:,.2f} m</div>
        </div>
    """, unsafe_allow_html=True)

with col_k4:
    st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">🔧 Tipos</div>
            <div class="metric-value">{kpis['tipos_count']}</div>
        </div>
    """, unsafe_allow_html=True)

with col_k5:
    if kpis['total_mp'] > 0:
        pct_rimado = (kpis['total_rimado'] / kpis['total_mp']) * 100
    else:
        pct_rimado = 0
    st.markdown(f"""
        <div class="metric-container">
            <div class="metric-label">📊 % Rimado</div>
            <div class="metric-value">{pct_rimado:.1f}%</div>
        </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ============================================
# PROCESAR Y MOSTRAR DATOS (CON HTML)
# ============================================

with st.spinner("Generando reporte..."):
    consumos_data = process_consumos(fecha_inicio, fecha_fin, año_seleccionado, mes_seleccionado)
    df_metros = process_metros(fecha_inicio, fecha_fin, año_seleccionado, mes_seleccionado)

# ============================================
# SECCIÓN 1: TABLA DE CONSUMOS CON HTML
# ============================================

st.header("📊 Consumos por Estado y Tipo Perforación")

if consumos_data:
    for estado in sorted(consumos_data.keys()):
        st.subheader(f"📌 Estado: {estado}")
        
        for tipo in sorted(consumos_data[estado].keys()):
            df = consumos_data[estado][tipo]
            
            if not df.empty:
                html_tabla = crear_tabla_html(df, titulo=f"🔧 {tipo}")
                st.markdown(html_tabla, unsafe_allow_html=True)
                st.markdown("---")
else:
    st.info("ℹ️ No hay datos de consumo para los filtros seleccionados")

# ============================================
# SECCIÓN 2: RESUMEN DE METROS CON HTML
# ============================================

st.header("📏 Resumen de Metros por Tipo Perforación")

if not df_metros.empty:
    html_tabla = crear_tabla_html(df_metros, titulo="📊 Metros por Tipo Perforación")
    st.markdown(html_tabla, unsafe_allow_html=True)
    
    st.subheader("📊 Comparativa de Metros por Tipo")
    
    df_metros_melt = df_metros.melt(
        id_vars=['Tipo Perforación'],
        var_name='Tipo',
        value_name='Metros'
    )
    
    fig = px.bar(
        df_metros_melt,
        x='Tipo Perforación',
        y='Metros',
        color='Tipo',
        barmode='group',
        title="MP Total vs MP Rimado por Tipo Perforación",
        labels={'Metros': 'Metros (m)', 'Tipo Perforación': 'Tipo Perforación'},
        color_discrete_map={'MP Total': '#2c3e50', 'MP Rimado': '#e67e22'}
    )
    fig.update_layout(
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        height=400,
        font=dict(size=12, family='Segoe UI, Arial, sans-serif'),
        title_font=dict(size=16, family='Segoe UI, Arial, sans-serif')
    )
    fig.update_traces(textposition='outside', texttemplate='%{y:.1f}m')
    st.plotly_chart(fig, use_container_width=True)
    
else:
    st.info("ℹ️ No hay datos de metros para los filtros seleccionados")

st.markdown("---")

# ============================================
# SECCIÓN 3: GRÁFICOS ADICIONALES
# ============================================

st.header("📈 Análisis Adicional")

if consumos_data:
    datos_consumo = []
    for estado in consumos_data:
        for tipo in consumos_data[estado]:
            df = consumos_data[estado][tipo]
            if not df.empty and 'Total' in df.columns:
                total = df['Total'].sum()
                datos_consumo.append({
                    'Estado': estado,
                    'Tipo': tipo,
                    'Consumo': total
                })
    
    if datos_consumo:
        df_consumo = pd.DataFrame(datos_consumo)
        
        col1, col2 = st.columns(2)
        
        with col1:
            consumo_estado = df_consumo.groupby('Estado')['Consumo'].sum().reset_index()
            fig2 = px.pie(
                consumo_estado,
                values='Consumo',
                names='Estado',
                title="Distribución de Consumo por Estado",
                hole=0.3
            )
            fig2.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                height=350,
                font=dict(size=12)
            )
            st.plotly_chart(fig2, use_container_width=True)
        
        with col2:
            consumo_tipo = df_consumo.groupby('Tipo')['Consumo'].sum().reset_index()
            fig3 = px.pie(
                consumo_tipo,
                values='Consumo',
                names='Tipo',
                title="Distribución de Consumo por Tipo Perforación",
                hole=0.3
            )
            fig3.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                height=350,
                font=dict(size=12)
            )
            st.plotly_chart(fig3, use_container_width=True)

# ============================================
# SECCIÓN 4: TOP 10 CON HTML
# ============================================

if consumos_data:
    st.subheader("🏆 Top 10 Aceros más Consumidos")
    
    top_consumos = []
    for estado in consumos_data:
        for tipo in consumos_data[estado]:
            df = consumos_data[estado][tipo]
            if not df.empty:
                for _, row in df.iterrows():
                    top_consumos.append({
                        'Código': row['Codigo'],
                        'Descripción': row['Descripcion'][:40] + '...' if len(row['Descripcion']) > 40 else row['Descripcion'],
                        'Consumo': row['Total'],
                        'Estado': estado,
                        'Tipo': tipo
                    })
    
    if top_consumos:
        df_top = pd.DataFrame(top_consumos)
        df_top = df_top.sort_values('Consumo', ascending=False).head(10)
        
        html_tabla = crear_tabla_html(
            df_top[['Código', 'Descripción', 'Consumo', 'Estado', 'Tipo']],
            titulo="🏆 Top 10 Aceros más Consumidos"
        )
        st.markdown(html_tabla, unsafe_allow_html=True)


# ============================================
# SECCIÓN: RENDIMIENTO DE ACEROS
# ============================================

st.header("🏆 Rendimiento de Aceros")

with st.spinner("Procesando rendimiento de aceros..."):
    rendimiento_aceros = process_rendimiento_aceros(
        fecha_inicio, fecha_fin, año_seleccionado, mes_seleccionado, "TODAS"
    )

if rendimiento_aceros:
    for tipo, df in rendimiento_aceros.items():
        if not df.empty:
            st.markdown(f"### 📌 {tipo}")
            
            # Formatear columnas
            df_display = df.copy()
            df_display['Cantidad'] = df_display['Cantidad'].apply(lambda x: f"{x:,.0f}")
            df_display['Metros'] = df_display['Metros'].apply(lambda x: f"{x:,.2f}")
            df_display['Rendimiento'] = df_display['Rendimiento'].apply(lambda x: f"{x:,.2f}")
            df_display['Objetivo'] = df_display['Objetivo'].apply(lambda x: f"{x:,.2f}")
            df_display['Eficiencia'] = df_display['Eficiencia'].apply(lambda x: f"{x:.1f}%")
            
            html_rendimiento = crear_tabla_html(
                df_display,
                titulo=""
            )
            st.markdown(html_rendimiento, unsafe_allow_html=True)
            
            # Gráfico de eficiencia (opcional)
            if len(df) > 1:
                fig_rendimiento = px.bar(
                    df,
                    x='Familia',
                    y='Eficiencia',
                    color='Eficiencia',
                    color_continuous_scale='RdYlGn',
                    title=f"Eficiencia por Familia - {tipo}",
                    text=df['Eficiencia'].apply(lambda x: f"{x:.1f}%")
                )
                fig_rendimiento.update_traces(textposition='outside')
                fig_rendimiento.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    height=300,
                    margin=dict(l=0, r=0, t=40, b=0)
                )
                st.plotly_chart(fig_rendimiento, use_container_width=True)
            
            st.markdown("---")
else:
    st.info("ℹ️ No hay datos de rendimiento de aceros para los filtros seleccionados")


# ============================================
# PIE DE PÁGINA
# ============================================

st.markdown("---")
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    st.caption(f"📅 Reporte generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    st.caption(f"📊 Datos filtrados: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}")